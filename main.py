import os
import shutil
import csv
import pytesseract
from PIL import Image
from tkinter import filedialog, Tk
from openpyxl import Workbook

# Configure pytesseract to point to the installed Tesseract-OCR
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def process_text(text):
    words = text.lower().split()
    words = [word for word in words if word.isalnum()]
    unique_words = list(set(words))
    final_name = " ".join(unique_words)
    return final_name

def rename_gifs(gif_folder, save_folder):
    gif_files = [f for f in os.listdir(gif_folder) if f.endswith('.gif')]
    
    # Create necessary directories
    high_accuracy_dir = os.path.join(save_folder, "renamed", "high_accuracy")
    mid_accuracy_dir = os.path.join(save_folder, "renamed", "mid_accuracy")
    low_accuracy_dir = os.path.join(save_folder, "renamed", "low_accuracy")
    error_dir = os.path.join(save_folder, "renamed", "error")

    os.makedirs(high_accuracy_dir, exist_ok=True)
    os.makedirs(mid_accuracy_dir, exist_ok=True)
    os.makedirs(low_accuracy_dir, exist_ok=True)
    os.makedirs(error_dir, exist_ok=True)

    # Prepare master CSV
    master_csv = Workbook()
    high_sheet = master_csv.create_sheet("High Accuracy")
    mid_sheet = master_csv.create_sheet("Mid Accuracy")
    low_sheet = master_csv.create_sheet("Low Accuracy")
    error_sheet = master_csv.create_sheet("Errors")

    high_sheet.append(["Original Filename", "New Filename"])
    mid_sheet.append(["Original Filename", "New Filename"])
    low_sheet.append(["Original Filename", "New Filename"])
    error_sheet.append(["Original Filename", "Error Message"])

    for gif_file in gif_files:
        try:
            file_path = os.path.join(gif_folder, gif_file)
            image = Image.open(file_path)
            extracted_text = pytesseract.image_to_string(image)
            new_name = process_text(extracted_text)

            if not new_name or new_name == "untitled":
                shutil.copy(file_path, os.path.join(error_dir, gif_file))
                error_sheet.append([gif_file, "Renamed as 'untitled.gif'"])
                continue

            new_file_name = new_name + ".gif"
            accuracy = len(extracted_text) / (len(image.getdata()) if len(image.getdata()) > 0 else 1)

            if accuracy >= 0.9:
                dest_dir = high_accuracy_dir
                high_sheet.append([gif_file, new_file_name])
            elif accuracy >= 0.5:
                dest_dir = mid_accuracy_dir
                mid_sheet.append([gif_file, new_file_name])
            else:
                dest_dir = low_accuracy_dir
                low_sheet.append([gif_file, new_file_name])

            new_file_path = os.path.join(dest_dir, new_file_name)
            shutil.copy(file_path, new_file_path)

        except Exception as e:
            shutil.copy(file_path, os.path.join(error_dir, gif_file))
            error_sheet.append([gif_file, str(e)])

    # Save master CSV
    master_csv.save(os.path.join(save_folder, "renamed", "master.csv"))

# GUI for folder selection
if __name__ == "__main__":
    root = Tk()
    root.withdraw()  # Hide the main window
    gif_folder = filedialog.askdirectory(title="Select the GIF folder")
    save_folder = filedialog.askdirectory(title="Select the save folder")
    rename_gifs(gif_folder, save_folder)
    print("GIF renaming process completed.")
